from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def _norm_header(value: Any) -> str:
    return str(value or "").strip()


def _guess_type_from_sample(value: Any) -> str:
    if value is None or value == "":
        return "string"

    if isinstance(value, bool):
        return "string"

    if isinstance(value, int):
        return "int"

    if isinstance(value, float):
        if value.is_integer():
            return "int"
        return "string"

    cls_name = value.__class__.__name__.lower()
    if "date" in cls_name or "datetime" in cls_name:
        return "date"

    s = str(value).strip()
    s_lower = s.lower()

    if s_lower.startswith("http://") or s_lower.startswith("https://") or s_lower.startswith("www."):
        return "string"

    if s.isdigit():
        return "int"

    slash_parts = s.split("/")
    dash_parts = s.split("-")

    if len(slash_parts) == 3 and all(part.strip().isdigit() for part in slash_parts):
        return "date"

    if len(dash_parts) == 3 and all(part.strip().isdigit() for part in dash_parts):
        return "date"

    return "string"


def _cell_in_same_data_column(cell_range: str, col_letter: str) -> bool:
    try:
        parts = cell_range.split(":")
        start = parts[0]
        end = parts[-1]

        start_letters = "".join(ch for ch in start if ch.isalpha()).upper()
        end_letters = "".join(ch for ch in end if ch.isalpha()).upper()

        if not start_letters:
            return False
        if not end_letters:
            end_letters = start_letters

        return start_letters <= col_letter <= end_letters
    except Exception:
        return False


def _read_values_from_range_ref(workbook, ref: str, debug: bool = False) -> List[str]:
    if debug:
        print("READ RANGE REF =", ref)

    if "!" in ref:
        sheet_name, range_ref = ref.split("!", 1)
        sheet_name = sheet_name.strip("'")
        ws = workbook[sheet_name]
    else:
        ws = workbook.active
        range_ref = ref

    values: List[str] = []
    for row in ws[range_ref]:
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                values.append(str(cell.value).strip())
    return values


def _extract_enum_from_validation(
    workbook,
    ws,
    col_idx: int,
    sample_row: int,
    debug: bool = False,
) -> List[str]:
    col_letter = get_column_letter(col_idx)
    target_coord = f"{col_letter}{sample_row}"

    if debug:
        print(f"\n--- DEBUG ENUM COLONNA {col_letter} ({target_coord}) ---")
        print(f"Tot dataValidations: {len(ws.data_validations.dataValidation)}")

    for dv in ws.data_validations.dataValidation:
        if not isinstance(dv, DataValidation):
            continue

        if debug:
            print("type =", dv.type, "| formula1 =", dv.formula1, "| sqref =", dv.sqref)

        if dv.type != "list":
            continue

        applies = False

        for rng in str(dv.sqref).split():
            if _cell_in_same_data_column(rng, col_letter):
                applies = True
                break

        if not applies:
            try:
                if target_coord in dv.cells:
                    applies = True
            except Exception:
                pass

        if not applies:
            continue

        formula1 = dv.formula1
        if not formula1:
            continue

        formula1 = str(formula1).strip()

        if debug:
            print("MATCH validation list on column", col_letter)
            print("formula1 raw =", formula1)

        if formula1.startswith('"') and formula1.endswith('"'):
            raw = formula1[1:-1]
            values = [x.strip() for x in raw.split(",") if x.strip()]
            if debug:
                print("enum explicit =", values)
            return values

        ref = formula1[1:] if formula1.startswith("=") else formula1

        if "!" in ref or "$" in ref or ":" in ref:
            try:
                values = _read_values_from_range_ref(workbook, ref, debug=debug)
                if debug:
                    print("enum from range =", values)
                return values
            except Exception as exc:
                if debug:
                    print("range parse failed:", exc)
                return []

    if debug:
        print("No enum validation found for column", col_letter)

    return []


def _is_formula_cell(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _build_field_def(
    name: str,
    sample_value: Any,
    has_formula: bool,
    formula_source: str | None,
    formula_anchor_row: int | None,
    enum_values: List[str],
) -> Dict[str, Any]:
    base = {
        "name": name,
        "required": False,
        "visibleInViewer": True,
        "enumValues": enum_values or [],
        "locked": False,
        "formulaSource": None,
        "formulaAnchorRow": None,
        "formulaMode": None,
        "enumStyles": {},
    }

    if name == "id":
        return {
            **base,
            "type": "int",
            "computed": True,
            "visibleInForm": False,
            "locked": True,
        }

    if has_formula:
        return {
            **base,
            "type": "computed",
            "computed": True,
            "visibleInForm": False,
            "locked": True,
            "formulaSource": formula_source,
            "formulaAnchorRow": formula_anchor_row,
            "formulaMode": "incremental_copy",
        }

    field_type = "enum" if enum_values else _guess_type_from_sample(sample_value)

    return {
        **base,
        "type": field_type,
        "computed": False,
        "visibleInForm": True,
    }


def _build_constraint(field: Dict[str, Any]) -> Dict[str, Any]:
    name = field["name"]
    ftype = field["type"]

    if name == "id":
        return {
            "type": "int",
            "required": False,
            "min": 0,
        }

    if ftype == "int":
        return {
            "type": "int",
            "required": bool(field["required"]),
            "min": 0,
        }

    if ftype == "date":
        return {
            "type": "date",
            "required": bool(field["required"]),
            "format": "dd/mm/yyyy",
        }

    if ftype == "enum":
        return {
            "type": "enum",
            "required": bool(field["required"]),
        }

    if ftype == "computed":
        return {
            "type": "computed",
            "required": False,
        }

    return {
        "type": "string",
        "required": bool(field["required"]),
        "maxLen": 255,
    }


def import_schema_from_xlsx(
    xlsx_path: str | Path,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    sample_row: int = 2,
    debug: bool = False,
) -> Dict[str, Any]:
    xlsx_path = Path(xlsx_path)

    wb = load_workbook(filename=xlsx_path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers: List[str] = []
    fields: List[Dict[str, Any]] = []

    col_idx = 1
    while True:
        cell_value = ws.cell(row=header_row, column=col_idx).value
        header = _norm_header(cell_value)

        if not header:
            break

        sample_cell = ws.cell(row=sample_row, column=col_idx)
        sample_value = sample_cell.value
        has_formula = _is_formula_cell(sample_cell)
        formula_source = str(sample_cell.value).strip() if has_formula else None
        formula_anchor_row = sample_row if has_formula else None

        enum_values = _extract_enum_from_validation(
            workbook=wb,
            ws=ws,
            col_idx=col_idx,
            sample_row=sample_row,
            debug=debug,
        )

        field = _build_field_def(
            name=header,
            sample_value=sample_value,
            has_formula=has_formula,
            formula_source=formula_source,
            formula_anchor_row=formula_anchor_row,
            enum_values=enum_values,
        )

        headers.append(header)
        fields.append(field)

        if debug:
            print(
                f"FIELD {header}: "
                f"sample={sample_value!r}, "
                f"formula={has_formula}, "
                f"formula_source={formula_source!r}, "
                f"enum={enum_values}, "
                f"type={field['type']}"
            )

        col_idx += 1

    visible_in_form = [f["name"] for f in fields if bool(f["visibleInForm"])]
    visible_in_viewer = [f["name"] for f in fields if bool(f["visibleInViewer"])]
    computed = [f["name"] for f in fields if bool(f["computed"])]
    required_on_insert = [
        f["name"] for f in fields
        if bool(f["required"]) and not bool(f["computed"])
    ]
    optional_on_insert = [
        f["name"] for f in fields
        if not bool(f["required"]) and not bool(f["computed"])
    ]

    constraints = {f["name"]: _build_constraint(f) for f in fields}
    enums = {f["name"]: f["enumValues"] for f in fields if f["enumValues"]}

    return {
        "headers": headers,
        "requiredOnInsert": required_on_insert,
        "optionalOnInsert": optional_on_insert,
        "computed": computed,
        "visibleInForm": visible_in_form,
        "visibleInViewer": visible_in_viewer,
        "enums": enums,
        "constraints": constraints,
        "fields": fields,
        "enumStyles": {},
    }


def save_schema_from_xlsx(
    xlsx_path: str | Path,
    output_json_path: str | Path | None,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    sample_row: int = 2,
    debug: bool = False,
) -> Dict[str, Any]:
    schema = import_schema_from_xlsx(
        xlsx_path=xlsx_path,
        sheet_name=sheet_name,
        header_row=header_row,
        sample_row=sample_row,
        debug=debug,
    )

    if output_json_path is not None:
        output_json_path = Path(output_json_path)
        output_json_path.parent.mkdir(parents=True, exist_ok=True)
        output_json_path.write_text(
            json.dumps(schema, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return schema