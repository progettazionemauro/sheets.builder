from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


def _norm_header(value: Any) -> str:
    return str(value or "").strip()


def _guess_type_from_sample(value: Any) -> str:
    if value is None or value == "":
        return "string"

    if isinstance(value, bool):
        return "string"

    if isinstance(value, int):
        return "int"

    if isinstance(value, float):
        if value.is_integer():
            return "int"
        return "string"

    cls_name = value.__class__.__name__.lower()
    if "date" in cls_name or "datetime" in cls_name:
        return "date"

    s = str(value).strip()
    s_lower = s.lower()

    if s_lower.startswith("http://") or s_lower.startswith("https://") or s_lower.startswith("www."):
        return "string"

    if s.isdigit():
        return "int"

    slash_parts = s.split("/")
    dash_parts = s.split("-")

    if len(slash_parts) == 3 and all(part.strip().isdigit() for part in slash_parts):
        return "date"

    if len(dash_parts) == 3 and all(part.strip().isdigit() for part in dash_parts):
        return "date"

    return "string"


def _cell_in_same_data_column(cell_range: str, col_letter: str) -> bool:
    try:
        parts = cell_range.split(":")
        start = parts[0]
        end = parts[-1]

        start_letters = "".join(ch for ch in start if ch.isalpha()).upper()
        end_letters = "".join(ch for ch in end if ch.isalpha()).upper()

        if not start_letters:
            return False
        if not end_letters:
            end_letters = start_letters

        return start_letters <= col_letter <= end_letters
    except Exception:
        return False


def _read_values_from_range_ref(workbook, ref: str, debug: bool = False) -> List[str]:
    if debug:
        print("READ RANGE REF =", ref)

    if "!" in ref:
        sheet_name, range_ref = ref.split("!", 1)
        sheet_name = sheet_name.strip("'")
        ws = workbook[sheet_name]
    else:
        ws = workbook.active
        range_ref = ref

    values: List[str] = []
    for row in ws[range_ref]:
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                values.append(str(cell.value).strip())
    return values


def _extract_enum_from_validation(
    workbook,
    ws,
    col_idx: int,
    sample_row: int,
    debug: bool = False,
) -> List[str]:
    col_letter = get_column_letter(col_idx)
    target_coord = f"{col_letter}{sample_row}"

    if debug:
        print(f"\n--- DEBUG ENUM COLONNA {col_letter} ({target_coord}) ---")
        print(f"Tot dataValidations: {len(ws.data_validations.dataValidation)}")

    for dv in ws.data_validations.dataValidation:
        if not isinstance(dv, DataValidation):
            continue

        if debug:
            print("type =", dv.type, "| formula1 =", dv.formula1, "| sqref =", dv.sqref)

        if dv.type != "list":
            continue

        applies = False

        for rng in str(dv.sqref).split():
            if _cell_in_same_data_column(rng, col_letter):
                applies = True
                break

        if not applies:
            try:
                if target_coord in dv.cells:
                    applies = True
            except Exception:
                pass

        if not applies:
            continue

        formula1 = dv.formula1
        if not formula1:
            continue

        formula1 = str(formula1).strip()

        if debug:
            print("MATCH validation list on column", col_letter)
            print("formula1 raw =", formula1)

        if formula1.startswith('"') and formula1.endswith('"'):
            raw = formula1[1:-1]
            values = [x.strip() for x in raw.split(",") if x.strip()]
            if debug:
                print("enum explicit =", values)
            return values

        ref = formula1[1:] if formula1.startswith("=") else formula1

        if "!" in ref or "$" in ref or ":" in ref:
            try:
                values = _read_values_from_range_ref(workbook, ref, debug=debug)
                if debug:
                    print("enum from range =", values)
                return values
            except Exception as exc:
                if debug:
                    print("range parse failed:", exc)
                return []

    if debug:
        print("No enum validation found for column", col_letter)

    return []


