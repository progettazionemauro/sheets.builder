from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


def _norm_header(value: Any) -> str:
    return str(value or "").strip()


def _guess_type_from_sample(value: Any) -> str:
    if value is None or value == "":
        return "string"

    if isinstance(value, bool):
        return "string"

    if isinstance(value, int):
        return "int"

    if isinstance(value, float):
        if value.is_integer():
            return "int"
        return "string"

    # openpyxl può restituire datetime/date; per semplicità trattiamo come date
    cls_name = value.__class__.__name__.lower()
    if "date" in cls_name or "datetime" in cls_name:
        return "date"

    s = str(value).strip()

    # euristica semplice per date tipo 14/01/1965 o 2026-03-28
    if "/" in s and len(s) >= 8:
        return "date"
    if "-" in s and len(s) >= 8:
        return "date"

    # numero intero in stringa
    if s.isdigit():
        return "int"

    return "string"


def _extract_enum_from_validation(
    ws,
    header_row: int,
    sample_row: int,
    col_idx: int,
) -> List[str]:
    """
    Cerca una data validation applicata alla cella della riga sample_row / col_idx.
    Se trova una lista esplicita o una lista da range, restituisce enumValues.
    Altrimenti restituisce [].
    """
    target_cell = ws.cell(row=sample_row, column=col_idx)
    target_coord = target_cell.coordinate

    for dv in ws.data_validations.dataValidation:
        if not isinstance(dv, DataValidation):
            continue

        if target_coord not in dv.cells:
            continue

        if dv.type != "list":
            return []

        formula1 = dv.formula1
        if not formula1:
            return []

        formula1 = str(formula1).strip()

        # Caso 1: lista esplicita tipo "A,B,C"
        if formula1.startswith('"') and formula1.endswith('"'):
            raw = formula1[1:-1]
            values = [x.strip() for x in raw.split(",") if x.strip()]
            return values

        # Caso 2: riferimento a range tipo =Foglio2!$A$1:$A$5
        if formula1.startswith("="):
            ref = formula1[1:]
            try:
                values = _read_values_from_range_ref(ws.parent, ref)
                return [v for v in values if str(v).strip()]
            except Exception:
                return []

    return []


def _read_values_from_range_ref(workbook, ref: str) -> List[str]:
    """
    Legge valori da un riferimento tipo:
    - Sheet2!$A$1:$A$5
    - $A$1:$A$5  (stesso foglio, se mai capitasse)
    """
    if "!" in ref:
        sheet_name, range_ref = ref.split("!", 1)
        sheet_name = sheet_name.strip("'")
        ws = workbook[sheet_name]
    else:
        # fallback poco usato
        ws = workbook.active
        range_ref = ref

    values: List[str] = []
    for row in ws[range_ref]:
        for cell in row:
            if cell.value is not None and str(cell.value).strip():
                values.append(str(cell.value).strip())
    return values


def _is_formula_cell(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _build_field_def(
    name: str,
    sample_value: Any,
    has_formula: bool,
    enum_values: List[str],
) -> Dict[str, Any]:
    if name == "id":
        return {
            "name": "id",
            "type": "int",
            "required": False,
            "computed": True,
            "visibleInForm": False,
            "visibleInViewer": True,
            "enumValues": [],
            "locked": True,
        }

    if has_formula:
        return {
            "name": name,
            "type": "computed",
            "required": False,
            "computed": True,
            "visibleInForm": False,
            "visibleInViewer": True,
            "enumValues": [],
            "locked": True,
        }

    if enum_values:
        field_type = "enum"
    else:
        field_type = _guess_type_from_sample(sample_value)

    return {
        "name": name,
        "type": field_type,
        "required": False,
        "computed": False,
        "visibleInForm": True,
        "visibleInViewer": True,
        "enumValues": enum_values,
        "locked": False,
    }


def _build_constraint(field: Dict[str, Any]) -> Dict[str, Any]:
    name = field["name"]
    ftype = field["type"]

    if name == "id":
        return {
            "type": "int",
            "required": False,
            "min": 0,
        }

    if ftype == "int":
        return {
            "type": "int",
            "required": bool(field["required"]),
            "min": 0,
        }

    if ftype == "date":
        return {
            "type": "date",
            "required": bool(field["required"]),
            "format": "dd/mm/yyyy",
        }

    if ftype == "enum":
        return {
            "type": "enum",
            "required": bool(field["required"]),
        }

    if ftype == "computed":
        return {
            "type": "computed",
            "required": False,
        }

    return {
        "type": "string",
        "required": bool(field["required"]),
        "maxLen": 255,
    }


def import_schema_from_xlsx(
    xlsx_path: str | Path,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    sample_row: int = 2,
) -> Dict[str, Any]:
    xlsx_path = Path(xlsx_path)

    wb = load_workbook(filename=xlsx_path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers: List[str] = []
    fields: List[Dict[str, Any]] = []

    # Leggiamo la riga header finché ci sono celle non vuote
    col_idx = 1
    while True:
        cell_value = ws.cell(row=header_row, column=col_idx).value
        header = _norm_header(cell_value)

        if not header:
            break

        sample_cell = ws.cell(row=sample_row, column=col_idx)
        sample_value = sample_cell.value
        has_formula = _is_formula_cell(sample_cell)
        enum_values = _extract_enum_from_validation(ws, header_row, sample_row, col_idx)

        field = _build_field_def(
            name=header,
            sample_value=sample_value,
            has_formula=has_formula,
            enum_values=enum_values,
        )

        headers.append(header)
        fields.append(field)

        col_idx += 1

    visible_in_form = [
        f["name"] for f in fields
        if bool(f["visibleInForm"])
    ]

    visible_in_viewer = [
        f["name"] for f in fields
        if bool(f["visibleInViewer"])
    ]

    computed = [
        f["name"] for f in fields
        if bool(f["computed"])
    ]

    required_on_insert = [
        f["name"] for f in fields
        if bool(f["required"]) and not bool(f["computed"])
    ]

    optional_on_insert = [
        f["name"] for f in fields
        if not bool(f["required"]) and not bool(f["computed"])
    ]

    constraints = {
        f["name"]: _build_constraint(f)
        for f in fields
    }

    enums = {
        f["name"]: f["enumValues"]
        for f in fields
        if f["enumValues"]
    }

    return {
        "headers": headers,
        "requiredOnInsert": required_on_insert,
        "optionalOnInsert": optional_on_insert,
        "computed": computed,
        "visibleInForm": visible_in_form,
        "visibleInViewer": visible_in_viewer,
        "enums": enums,
        "constraints": constraints,
        "fields": fields,
    }


def save_schema_from_xlsx(
    xlsx_path: str | Path,
    output_json_path: str | Path,
    sheet_name: Optional[str] = None,
    header_row: int = 1,
    sample_row: int = 2,
) -> Dict[str, Any]:
    schema = import_schema_from_xlsx(
        xlsx_path=xlsx_path,
        sheet_name=sheet_name,
        header_row=header_row,
        sample_row=sample_row,
    )

    output_json_path = Path(output_json_path)
    output_json_path.parent.mkdir(parents=True, exist_ok=True)
    output_json_path.write_text(
        json.dumps(schema, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return schema
